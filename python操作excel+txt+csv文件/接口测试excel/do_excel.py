"""
读取测试用例 完成excel中用例的读写统计
"""
from  openpyxl import load_workbook
from ddt import ddt,data,unpack

class doExcel:
    # 类初始化的时候必须传入两个参数
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name
    def get_header(self):
        """获取第一行的标题"""
        wb=load_workbook(self.file_name)# 获取excel
        sheet=wb[self.sheet_name]# 获取excel的sheet
        headers=[]
        for i in range(1,sheet.max_column+1):
            headers.append(sheet.cell(1,i).value)
        return headers
    def get_data(self):
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]# 这个地方不能写成真字符串
        test_data=[]
        header=self.get_header()
        for i in range(1,sheet.max_row+1):# 获取最大的行数
            sub={}
            for j in range(1,sheet.max_column+1):
                sub["method"]=sheet.cell(i,1).value
                sub['url']=sheet.cell(i,2).value
                sub['data']=sheet.cell(i,3).value
                sub['expected']=sheet.cell(i,4).value
            test_data.append(sub)
        return test_data




