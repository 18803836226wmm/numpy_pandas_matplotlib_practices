from openpyxl import load_workbook
import read_config
import project_path

class DoExcel:
    @staticmethod
    def get_data(file_name):
        wb=load_workbook(file_name)
        """注意读取配置文件中的数据要eval一下，读出的格式是str """
        mode=eval(read_config.ReadConfig(project_path.case_confg_path,'MODE','mode'))
        test_data=[] #数组放在这个位置上
        # 遍历这个存在配置文件的字典
        for key in mode:
            sheet=wb[key]
            # 如果model是all，则需要读取所有的数据，配置文件的参数一定是字符串
            if mode[key]=='all':
                for i in range(2,sheet.max_row+1):
                    row_data = {}  # 定义一个字典 把读出的数据分别存储到表格里面
                    row_data['case_id'] = sheet.cell(i,1).value
                    row_data['url']=sheet.cell(i,2).value
                    row_data['data']=sheet.cell(i,3).value
                    row_data['title']=sheet.cell(i,4).value
                    row_data['http_method']=sheet.cell(i,5).value
                    # 增加一个字段用来控制测试用例结果写入
                    row_data['sheet_name']=key
                    # 要注意 从表格里读取的数据是什么类型的 实际请求的接口就需要什么类型的参数，要记得eval一下
                    test_data.append(row_data)
            else:
                for case_id in mode[key]:
                    row_data = {}
                    row_data["case_id"] = sheet.cell(case_id + 1, 1).value  # 行号
                    row_data["url"] = sheet.cell(case_id + 1, 2).value
                    row_data["data"] = sheet.cell(case_id + 1, 3).value
                    row_data["title"] = sheet.cell(case_id + 1, 4).value
                    row_data["http_method"] = sheet.cell(case_id + 1, 5).value
                    row_data["sheet_name"] = key
                    test_data.append(row_data)
        return test_data
        # 专门用来写入数据
    @staticmethod
    def write_back(file_name,sheet_name,i,result,TestResult):
        wb=load_workbook(file_name)
        sheet=wb[sheet_name]
        sheet.cell(i,6).value=result
        sheet.cell(i,7).value=TestResult
        # 写完数据后保存
        wb.save(file_name)










